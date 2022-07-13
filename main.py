import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font


def get_items():
    try:
        wb = load_workbook('FWUO.xlsx')
        ws = wb.active
        if ws['Z1'].value != int(datetime.now().strftime('%H')):
            create_file()
        else:
            print('Обновлять можно раз в час.')
    except IOError:
        create_file()
    import time
    time.sleep(5)


def create_file():
    url = 'https://www.fwuo.ru/vendors/?query=*&vendor=all'
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    results = soup.find_all('td')
    results.pop(0)
    results.pop(0)
    results.pop(0)
    results.pop(0)
    count = 0
    counter = 1
    id = 0
    trade_list = ()
    all_list = list()
    for result in results:
        if count == 0:
            trade_list = {'id': int(id), 'Vendor': result.text}
        if count == 1:
            trade_list['Name'] = result.text
        if count == 2:
            trade_list['Count'] = int(result.text)
        if count == 3:
            trade_list['Price'] = int(result.text)
            count = -1
        id += 0.25
        count += 1
        counter += 1
        if counter == 4:
            all_list.append(trade_list)
            counter = 0
    list_id = []
    list_vendor = []
    list_name = []
    list_count = []
    list_price = []
    sorted_list = sorted(all_list, key=lambda d: d['Price'])
    for i in range(0, len(sorted_list)):
        list_id.append(sorted_list[i].get('id'))
        list_vendor.append(sorted_list[i].get('Vendor'))
        list_name.append(sorted_list[i].get('Name'))
        list_count.append(sorted_list[i].get('Count'))
        list_price.append(sorted_list[i].get('Price'))
    data = {'Имя Вендора': list_vendor, 'Название вещи': list_name, 'Количество': list_count,
            'Цена': list_price}
    df = pd.DataFrame(data)
    df.to_excel('FWUO.xlsx')
    wb = load_workbook('FWUO.xlsx')
    ws = wb.active
    ws.insert_rows(0, 1)
    ws['B1'] = datetime.now().strftime('%d/%m/%y %H:%M')
    ws['Z1'] = int(datetime.now().strftime('%H'))
    ws.auto_filter.ref = "B2:E2"
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws['Z1'].font = Font(color="FFFFFF")
    wb.save('FWUO.xlsx')
    print('Данные успешно добавлены.')

get_items()

# pip download --no-deps pyinstaller==4.5
# pip install pyinstaller-4.5-py3-none-win_amd64.whl
# pyinstaller --onefile main.py --exclude-module _bootlocale
